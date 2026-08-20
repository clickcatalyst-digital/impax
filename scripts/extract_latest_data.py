import openpyxl, json

path = "scripts/26.1.26-1-latest.xlsm"
wb_v = openpyxl.load_workbook(path, data_only=True, keep_vba=True)

def safe(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return v

def is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

# STOCK sheet -> use current "STOCK" (col K) as opening balance if present
# and numeric, else fall back to BALANCE (col H) if numeric, else 0.
# (Same logic as the original scripts/extract_seed_data.py.)
ws = wb_v['STOCK']
stock_items = []
skipped = []
flagged = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    a,b,c,d,e,f,g,h,i,j,k,l = [safe(row[idx].value) for idx in range(12)]

    if not b or (not a and not c and not d):
        if b:
            skipped.append({"row": row[0].row, "item": b, "reason": "no group/data/type - looks like a stray row"})
        continue

    # K (STOCK) is normally =BALANCE+IN2-OUT3 (verified against every row with
    # a working formula - zero mismatches). When K itself is broken (e.g. #REF!
    # from a deleted cell), recompute it from BALANCE+IN2-OUT3 instead of just
    # falling back to BALANCE alone, which would silently drop the IN2/OUT3 leg.
    if is_number(k):
        opening = k
    elif k is None:
        # No K formula at all for this row (never had a stage-2 IN2/OUT3
        # entry) - BALANCE is genuinely the current stock, nothing broken.
        opening = h if is_number(h) else 0
    elif is_number(h) and is_number(i) and is_number(j):
        opening = h + i - j
        flagged.append({"row": row[0].row, "item": b, "reason": f"K column had a broken formula ({k!r}) - recomputed as BALANCE+IN2-OUT3 = {opening}"})
    elif is_number(h):
        opening = h
        flagged.append({"row": row[0].row, "item": b, "reason": f"K column had a broken formula ({k!r}) and IN2/OUT3 missing - used BALANCE column instead"})
    else:
        opening = 0
        flagged.append({"row": row[0].row, "item": b, "reason": f"no valid numeric stock found (K={k!r}, H={h!r}) - set to 0"})

    t_val = safe(row[19].value) if len(row) > 19 else None
    t_val = t_val if is_number(t_val) else 0
    price = l if is_number(l) else 0
    stock_items.append({
        "group": a, "item": b, "data": c, "type": d,
        "opening_balance": opening, "price": price, "min_qty": t_val
    })

# BLANK PCB sheet
ws2 = wb_v['BLANK PCB']
pcb_items = []
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    a,b,c,d,e,f,g,h,i,j = [safe(row[idx].value) for idx in range(10)]
    if not b:
        continue
    opening = f if is_number(f) else 0
    pcb_items.append({
        "pcb_number": a, "name": b, "opening_balance": opening,
        "price": g if is_number(g) else 0, "min_qty": i if is_number(i) else 0
    })

print("STOCK items:", len(stock_items))
print("PCB items:", len(pcb_items))
print()
print("Skipped stray rows:", len(skipped))
for s in skipped:
    print(" ", s)
print("Flagged rows (fallback/zeroed values):", len(flagged))
for fl in flagged:
    print(" ", fl)

with open("scripts/seed-data-latest.json", "w") as f:
    json.dump({"stock": stock_items, "pcb": pcb_items}, f, indent=2)
